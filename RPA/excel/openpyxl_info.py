from os import minor
from openpyxl import Workbook
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.drawing.image import Image
from random import *

wb = Workbook()
ws = wb.active


ws.append(["번호", "영어", "수학"])
for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"]
# print(col_B)

# for cell in col_B:
#     print(cell.value)

col_range = ws["B:C"]
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

row_title = ws[1]
# for cell in row_title:
# print(cell.value)

row_range = ws[2:6]
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()

row_range = ws[2:ws.max_row]
for rows in row_range:
    for cell in rows:
        # print(cell.value, end=" ")
        # print(cell.coordinate, end=" ")
        xy = coordinate_from_string(cell.coordinate)
        # print(xy, end=" ")
    #     print(xy[0], end=" ")
    #     print(xy[1], end=" ")
    # print()

# print(tuple(ws.rows))
# for row in tuple(ws.rows):
#     print(row[1].value)
# print(tuple(ws.columns))
# for column in tuple(ws.columns):
#     print(column[0].value)

# for row in ws.iter_rows():
#     print(row[1].value)

# for row in ws.iter_cols():
#     print(row[0].value)

# for row in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3):
#     print(row[0].value, row[1].value)

# for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
#     print(col[0].value, col[1].value)

a1 = ws["A1"]
b1 = ws["B1"]
c1 = ws["C1"]

ws.column_dimensions["A"].width = 5
ws.row_dimensions[1].height = 50

a1.font = Font(color="FF0000", italic=True, bold=True)
b1.font = Font(color="CC33FF", name="Arial", strike=True)
c1.font = Font(color="0000FF", size=20, underline="single")

thin_border = Border(left=Side(style="thin"), right=Side(
    style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")

        if cell.column == 1:
            continue
        if isinstance(cell.value, int) and cell.value > 90:
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid")
            cell.font = Font(color="FF0000")

ws.freeze_panes = "B2"

bar_value = Reference(ws, min_row=2, max_row=11, min_col=2, max_col=3)
bar_chart = BarChart()
bar_chart.add_data(bar_value)
ws.add_chart(bar_chart, "E1")

line_value = Reference(ws, min_row=1, max_row=11, min_col=2, max_col=3)
line_chart = LineChart()
line_chart.add_data(line_value, titles_from_data=True)
line_chart.title = "성적표"
line_chart.style = 10
line_chart.y_axis.title = "점수"
line_chart.x_axis.title = "번호"
ws.add_chart(line_chart, "M1")

wb.save("chart.xlsx")


# ws.insert_rows(8, 5)
# ws.insert_cols(2)

# wb.save("chart_insert.xlsx")


# ws.delete_rows(5, 2)
# ws.delete_cols(2)

# wb.save("chart_delete.xlsx")

# ws.move_range("B1:C11", rows=0, cols=1)
# ws["B1"].value = "국어"
# ws.move_range("C1:C11", rows=5, cols=-1)
# wb.save("chart_move.xlsx")

ws.merge_cells("B12:D12")
ws["B2"].value = "Merged Cell"
wb.save("merge.xlsx")

ws.unmerge_cells("B12:D12")
wb.save("unmerge.xlsx")

img = Image("img.png")
ws.add_image(img, "C20")
wb.save("add_image.xlsx")
