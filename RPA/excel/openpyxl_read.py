from openpyxl import load_workbook
from random import *

wb = load_workbook("sample.xlsx")
ws = wb["Sheet"]

for x in range(1, 11):
    for y in range(1, 11):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()

wb.save("sample.xlsx")
wb.close()
